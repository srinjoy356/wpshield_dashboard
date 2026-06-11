import sys
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# BRAND DESIGN TOKENS (Hex colors without the # prefix for openpyxl)
DARK_TEAL = "0A6358"
LIGHT_TEAL_BG = "E6F4F1"
WHITE = "FFFFFF"
BLACK = "000000"
GRAY = "6B7280"
LIGHT_GRAY = "F9FAFB"

# SEVERITY COLORS
RED_TEXT = "DC2626"
ORANGE_TEXT = "EA580C"
GREEN_TEXT = "16A34A"

# Borders
THIN_BORDER_SIDE = Side(border_style="thin", color="D1D5DB")
GRID_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)

def apply_title_header(ws, title_text, col_span):
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    cell = ws.cell(row=1, column=1, value=title_text)
    cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 40
    
    # Subheader / Spacer row
    ws.row_dimensions[2].height = 15

def apply_table_headers(ws, headers, start_row=3):
    ws.row_dimensions[start_row].height = 26
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if header in ["Count", "Version", "Severity", "Risk Level", "Attack Count"] else "left", vertical="center")
        cell.border = GRID_BORDER

def autofit_columns(ws, start_row=3):
    for col in ws.columns:
        max_len = 10
        for cell in col:
            # Skip title row merged cells
            if cell.row < start_row:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 75)

def format_data_row(ws, row_idx, col_count):
    fill_color = LIGHT_TEAL_BG if row_idx % 2 == 1 else WHITE
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name="Calibri", size=10, color=BLACK)
        cell.fill = row_fill
        cell.border = GRID_BORDER
        # Default align
        cell.alignment = Alignment(vertical="center")

def format_attack_type(pattern):
    if not pattern:
        return ''
    mapping = {
        'sensitive_404': 'Sensitive File Probe',
        'sqli': 'SQL Injection',
        'xss': 'Cross-Site Scripting',
        'lfi': 'Local File Inclusion',
        'rce': 'Remote Code Execution',
        'xmlrpc': 'XML-RPC Abuse',
        'scanner_ua': 'Security Scanner',
        'wpscan_ua': 'WPScan Detected',
    }
    return mapping.get(pattern.lower(), pattern.replace('_', ' ').title())

def format_date(date_str):
    if not date_str:
        return '—'
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%b %d %Y, %I:%M %p')
    except:
        return date_str

def generate_excel(data):
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    company = data.get('company', {})
    maturity = data.get('maturity', {})
    stats = data.get('stats', {})
    score = maturity.get('score', 0)
    label = maturity.get('label', 'Needs Attention')
    
    # ─── TAB 1: SUMMARY ──────────────────────────────────────────
    ws = wb.create_sheet(title="Summary")
    apply_title_header(ws, "SECURITY REPORT SUMMARY", 5)
    
    # Metadata info block
    metadata = [
        ("Company Name", company.get('display_name', '')),
        ("Site URL", company.get('site_url', '')),
        ("Report Period", data.get('period', 'Last 30 Days')),
        ("Hardening Score", score),
        ("Hardening Level", label),
    ]
    
    # Custom block layout
    ws.cell(row=3, column=1, value="Report Details").font = Font(name="Calibri", size=12, bold=True, color=DARK_TEAL)
    ws.row_dimensions[3].height = 20
    
    for idx, (label_name, value) in enumerate(metadata, 4):
        ws.row_dimensions[idx].height = 22
        
        c_label = ws.cell(row=idx, column=1, value=label_name)
        c_label.font = Font(name="Calibri", size=10, bold=True, color=BLACK)
        c_label.fill = PatternFill(start_color=LIGHT_TEAL_BG, end_color=LIGHT_TEAL_BG, fill_type="solid")
        c_label.alignment = Alignment(vertical="center", indent=1)
        c_label.border = GRID_BORDER
        
        c_val = ws.cell(row=idx, column=2, value=value)
        c_val.font = Font(name="Calibri", size=10, bold=(label_name == "Hardening Score"), color=(DARK_TEAL if label_name == "Hardening Score" else BLACK))
        c_val.alignment = Alignment(vertical="center", indent=1)
        c_val.border = GRID_BORDER
        
        # Merge columns 2 to 5 for value layout aesthetics
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=5)
        for col_idx in range(3, 6):
            ws.cell(row=idx, column=col_idx).border = GRID_BORDER
            
    # Stats sub-block
    start_stats_row = 11
    ws.cell(row=start_stats_row, column=1, value="Core Security Metrics").font = Font(name="Calibri", size=12, bold=True, color=DARK_TEAL)
    ws.row_dimensions[start_stats_row].height = 20
    
    stats_headers = ["Metric", "Count"]
    apply_table_headers(ws, stats_headers, start_row=start_stats_row+1)
    
    stats_rows = [
        ("Total Attack Attempts", stats.get('totalAttacks', 0)),
        ("Total Login Attempts", stats.get('totalLogins', 0)),
        ("Total File Changes", stats.get('totalFileChanges', 0)),
        ("Open Security Alerts", stats.get('openAlerts', 0))
    ]
    
    for idx, (metric_name, val) in enumerate(stats_rows, start_stats_row+2):
        ws.row_dimensions[idx].height = 22
        format_data_row(ws, idx, 2)
        
        ws.cell(row=idx, column=1, value=metric_name).alignment = Alignment(vertical="center", indent=1)
        
        c_val = ws.cell(row=idx, column=2, value=val)
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        
        # Merge for visual alignment with metadata
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=5)
        for col_idx in range(3, 6):
            ws.cell(row=idx, column=col_idx).border = GRID_BORDER
            
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.views.sheetView[0].showGridLines = True
    
    # ─── TAB 2: ATTACKS ──────────────────────────────────────────
    ws = wb.create_sheet(title="Attacks")
    apply_title_header(ws, "TOP ATTACKING IP ADDRESSES", 3)
    apply_table_headers(ws, ["IP Address", "Attack Count", "Attack Type"])
    
    ips = data.get('topAttackingIps', [])
    for idx, ip in enumerate(ips, 4):
        ws.row_dimensions[idx].height = 22
        format_data_row(ws, idx, 3)
        
        c_ip = ws.cell(row=idx, column=1, value=ip.get('ip', ''))
        c_ip.font = Font(name="Calibri", size=10, bold=True, color=BLACK)
        c_ip.alignment = Alignment(vertical="center", indent=1)
        
        c_cnt = ws.cell(row=idx, column=2, value=ip.get('count', 0))
        c_cnt.alignment = Alignment(horizontal="center", vertical="center")
        
        c_type = ws.cell(row=idx, column=3, value=format_attack_type(ip.get('pattern_type', '')))
        c_type.alignment = Alignment(vertical="center", indent=1)
        
    autofit_columns(ws)
    ws.views.sheetView[0].showGridLines = True
    
    # ─── TAB 3: FILE CHANGES ──────────────────────────────────────
    ws = wb.create_sheet(title="File Changes")
    apply_title_header(ws, "RECENT FILE MODIFICATIONS", 3)
    apply_table_headers(ws, ["File Path", "Change Type", "Date"])
    
    files = data.get('recentFileChanges', [])
    for idx, f in enumerate(files, 4):
        ws.row_dimensions[idx].height = 22
        format_data_row(ws, idx, 3)
        
        c_path = ws.cell(row=idx, column=1, value=f.get('path', ''))
        c_path.alignment = Alignment(vertical="center", indent=1)
        
        change_type = f.get('event', '').replace('file_', '').title()
        c_change = ws.cell(row=idx, column=2, value=change_type)
        c_change.alignment = Alignment(horizontal="center", vertical="center")
        
        c_date = ws.cell(row=idx, column=3, value=format_date(f.get('occurred_at', '')))
        c_date.alignment = Alignment(vertical="center", indent=1)
        
    autofit_columns(ws)
    ws.views.sheetView[0].showGridLines = True
    
    # ─── TAB 4: VULNERABLE PLUGINS ───────────────────────────────
    ws = wb.create_sheet(title="Vulnerable Plugins")
    apply_title_header(ws, "VULNERABLE PLUGINS DETECTED", 5)
    apply_table_headers(ws, ["Plugin Name", "Version", "CVE ID", "Severity", "Fixed In"])
    
    vulns = data.get('vulnerablePlugins', [])
    for idx, p in enumerate(vulns, 4):
        ws.row_dimensions[idx].height = 22
        format_data_row(ws, idx, 5)
        
        ws.cell(row=idx, column=1, value=p.get('plugin_name', '')).alignment = Alignment(vertical="center", indent=1)
        
        c_ver = ws.cell(row=idx, column=2, value=p.get('plugin_version', ''))
        c_ver.alignment = Alignment(horizontal="center", vertical="center")
        
        c_cve = ws.cell(row=idx, column=3, value=p.get('cve_id', '—'))
        c_cve.alignment = Alignment(horizontal="center", vertical="center")
        
        severity_label = p.get('severity', '').upper()
        c_sev = ws.cell(row=idx, column=4, value=severity_label)
        c_sev.alignment = Alignment(horizontal="center", vertical="center")
        c_sev.font = Font(name="Calibri", size=10, bold=True, color=(RED_TEXT if "HIGH" in severity_label or "CRITICAL" in severity_label else ORANGE_TEXT))
        
        ws.cell(row=idx, column=5, value=p.get('fixed_in', 'Unpatched')).alignment = Alignment(vertical="center", indent=1)
        
    autofit_columns(ws)
    ws.views.sheetView[0].showGridLines = True
    
    # ─── TAB 5: ACTION ITEMS ──────────────────────────────────────
    ws = wb.create_sheet(title="Action Items")
    apply_title_header(ws, "ACTION ITEMS REQUIRED", 3)
    apply_table_headers(ws, ["Check Name", "Risk Level", "Recommendation"])
    
    FAIL_NAMES = {
        'No Vulnerable Plugins': 'Vulnerable Plugins Detected',
        'No High Open Alerts': 'Too Many High Severity Alerts',
        'No Critical Open Alerts': 'Critical Alerts Unresolved',
        'No Recent File Modification Alerts': 'Unexpected File Modifications',
        'Uptime Healthy': 'Site Offline or Unreachable',
        'Plugin Heartbeat Recent': 'Plugin Not Reporting Data',
        'HTTPS Enforced': 'HTTPS Not Configured',
    }
    
    failed = data.get('failedChecks', [])
    for idx, c in enumerate(failed, 4):
        ws.row_dimensions[idx].height = 24
        format_data_row(ws, idx, 3)
        
        display_name = FAIL_NAMES.get(c.get('check_name', ''), c.get('check_name', ''))
        ws.cell(row=idx, column=1, value=display_name).alignment = Alignment(vertical="center", indent=1)
        
        priority = c.get('priority', 'medium').upper()
        c_prior = ws.cell(row=idx, column=2, value=priority)
        c_prior.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color coding priority text
        if priority == "HIGH":
            c_prior.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)
        elif priority == "MEDIUM":
            c_prior.font = Font(name="Calibri", size=10, bold=True, color=ORANGE_TEXT)
        else:
            c_prior.font = Font(name="Calibri", size=10, bold=True, color=GREEN_TEXT)
            
        c_rec = ws.cell(row=idx, column=3, value=c.get('recommendation', ''))
        c_rec.alignment = Alignment(vertical="center", indent=1)
        
    autofit_columns(ws)
    ws.views.sheetView[0].showGridLines = True
    
    # Save to buffer and return bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

if __name__ == '__main__':
    # Stream input from stdin and write workbook to stdout
    data_json = json.loads(sys.stdin.read())
    excel_bytes = generate_excel(data_json)
    sys.stdout.buffer.write(excel_bytes)
